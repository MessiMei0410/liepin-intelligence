#!/usr/bin/env python3
"""Generate formatted Excel from cleaned recruitment data."""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import sys, json

# Read data from stdin as JSON
data = json.load(sys.stdin)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "招聘需求表"

# Styles
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="微软雅黑", size=10)
body_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

col_widths = [6, 12, 16, 24, 55, 55, 10, 10, 14]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

for row_idx, row_data in enumerate(data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        else:
            cell.font = body_font
            if col_idx in [1, 7, 8]:
                cell.alignment = center_align
            elif col_idx == 9:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = body_align
            if row_idx % 2 == 0:
                cell.fill = even_fill

ws.row_dimensions[1].height = 28
for r in range(2, len(data) + 1):
    ws.row_dimensions[r].height = 200
for tr in [2, 3, 8, 9, 10, 11, 12, 13]:
    if tr <= len(data):
        ws.row_dimensions[tr].height = 280

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{len(data)}"

sr = len(data) + 2
ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=6)
gap_total = sum(row[7] for row in data[1:])
sc = ws.cell(row=sr, column=1, value=f"合计：{len(data)-1} 个岗位，总缺口 {gap_total} 人")
sc.font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
sc.alignment = Alignment(horizontal="left", vertical="center")

out = sys.argv[1] if len(sys.argv) > 1 else "/Users/messi/Desktop/招聘需求表_整理版.xlsx"
wb.save(out)
print(f"✅ {out}  ({len(data)-1)}岗位, 缺口{gap_total}人)")
